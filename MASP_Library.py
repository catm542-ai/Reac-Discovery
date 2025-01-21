
"""
This library encompasses communication algorithms with respective devices and pre-designed processes for 
the automation and online monitoring of chemical reactions.
@author: Cristopher Tinajero
"""
# File: masf_library.py

import rseriesopc as rs
import matplotlib.pyplot as plt
import time
from datetime import datetime
import os
import numpy as np
import socket
import openpyxl
import csv
import pandas as pd
import cv2

def chronometer(tiempo_limite, a, o, setpressure):
    """
    This function enables a controlled stop prior to the continuation of the execution of the parent code.
    Additionally, it allows online monitoring and storage of pressure in the main R1 and secondary R2 
    modules of the system.

    Parameters:
        tiempo_limite (int): The waiting time in seconds.
        a (str): The name of the process being monitored.
        o (int): The number of the executed experiment (if necessary).
        setpressure (float): The target pressure for the system.

    Returns:
        None

    Notes:
        - Monitors pressure and logs data during the specified time period.
        - Saves recorded pressure data for future analysis.
    """
    ynn = []
    presion1n = []
    presion2n = []

    diferencia = 0
    for i in range(tiempo_limite):
        if diferencia <= tiempo_limite:
            diferencia += 1
            print(f"Elapsed time: {diferencia} seconds out of {tiempo_limite}, phase: {a}, Experiment: {o}")

            client = rs.RSeriesClient('opc.tcp://localhost:43344')
            try:
                isConnected = client.connect()
                presion = client.getRSeries().getManualControl().getModule(0)
                presion1 = presion.getPumpsPressures()
                presionpp = client.getRSeries().getManualControl().getModule(1)
                presion2 = presionpp.getPumpsPressures()

                if (round(presion1[0], 2)) >= (setpressure + 2.5) or (round(presion1[0], 2)) <= (setpressure - 2.5):
                    pass
            finally:
                if isConnected:
                    client.disconnect()

            ynn.append(int(diferencia))
            presion1n.append(round(presion1[0], 2))
            presion2n.append(round(presion2[0], 2))

        # Break the loop if the time limit is reached
        if diferencia >= tiempo_limite:
            break

        # Pause the program for 1 second
        time.sleep(1)

    print("Chronometer stopped after", tiempo_limite, "seconds")

    # Save pressure data to a text file
    datos = np.column_stack((ynn, presion1n, presion2n))
    np.savetxt(a + 'Experiment' + str(o) + '.txt', datos)

def sampleCollection(o, wt1, site, timer, setpressure):
    """
    Sends a command to the Autosampler to indicate the location and vial for collection.

    Parameters:
        o (int): The experiment number.
        wt1 (int): Waiting time before collection.
        site (int): The vial number for sample collection.
        timer (int): The collection duration (based on the required volume).
        setpressure (float): The target pressure for the system.

    Returns:
        None
    """
    class Valve:
        def __init__(self, stateSetter):
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)

    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        print("Sample Collection, Experiment #: " + str(o) + "\n")

        isConnected = client.connect()
        autosampler = client.getRSeries().getManualControl().getAutoSampler()
        autosampler.goToVialForCollection(site, timer)
        time.sleep(10)
        autosampler.setCollectionValveState(True)

        # Wait and monitor during sample collection
        a = "Sample Collection"
        chronometer(wt1, a, o, setpressure)

        time.sleep(5)
        autosampler.stopRoutines()
        time.sleep(5)
        autosampler.goHome()
        chronometer(70, "goHome", o, setpressure)
        autosampler.setCollectionValveState(False)
        time.sleep(5)
    finally:
        if isConnected:
            client.disconnect()

def finishingExperiment(o, setpressure):
    """
    Stops the manual execution of the system, returning the involved valves to their default position.

    Parameters:
        o (int): The experiment number.
        setpressure (float): Target pressure value.

    Returns:
        None
    """
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        print("Finalizing Experiment, Experiment #: " + str(o) + "\n")
        isConnected = client.connect()

        manual_control = client.getRSeries().getManualControl()
        for module_index in range(2):
            module = manual_control.getModule(module_index)
            module.getPumpA().setFlowRate(0)
            module.getPumpB().setFlowRate(0)
            module.getPumpA().setSRValveState(False)
            module.getPumpB().setSRValveState(False)

        manual_control.stopAll()
        chronometer(5, "Finalizing Experiment", o, setpressure)

    finally:
        if isConnected:
            client.disconnect()

def excelCapturing(rutaArchivo, o):
    """
    Reads experiment parameters from an Excel file for further use.

    Parameters:
        rutaArchivo (str): Path to the Excel file.
        o (int): Row offset for the experiment parameters.

    Returns:
        tuple: A set of extracted parameters including flow rates, waiting times, and reaction conditions.
    """
    ruta_archivo = rutaArchivo
    nombre_hoja = 'Hoja2'
    num_fila = o + 9

    archivo = openpyxl.load_workbook(ruta_archivo)
    hoja = archivo[nombre_hoja]

    fr1 = int(hoja.cell(row=num_fila, column=2).value)
    fr2 = int(hoja.cell(row=num_fila, column=3).value)
    fr3 = int(hoja.cell(row=num_fila, column=4).value)
    fr4 = int(hoja.cell(row=num_fila, column=5).value)
    wt1 = int(hoja.cell(row=num_fila, column=8).value)
    wt2 = int(hoja.cell(row=num_fila, column=9).value)
    wt3 = int(hoja.cell(row=num_fila, column=10).value)
    wt4 = int(hoja.cell(row=num_fila, column=11).value)
    wt5 = int(hoja.cell(row=num_fila, column=12).value)
    site = int(hoja.cell(row=num_fila, column=7).value)
    timer = int(hoja.cell(row=num_fila, column=10).value)
    reac = int(hoja.cell(row=num_fila, column=14).value)
    reacTemp = int(hoja.cell(row=num_fila, column=13).value)

    return fr1, fr2, fr3, fr4, wt1, wt2, wt3, wt4, wt5, reac, reacTemp, site, timer

def ReactorI_SolI(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the first reactor line and reagents from bottle I.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow Rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()
    
def ReactorII_SolII(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the first reactor line and reagents from bottle II.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()

def ReactorIII_SolIII(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the first reactor line and reagents from bottle III.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()

def ReactorIV_SolIV(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the second reactor line and reagents from bottle I.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(False)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()

def ReactorV_SolV(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the second reactor line and reagents from bottle II.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(False)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()

def ReactorVI_SolVI(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the second reactor line and reagents from bottle III.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(False)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()

def ReactorVII_SolVII(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the third  reactor line and reagents from bottle I.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(False)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()
            
def ReactorVIII_SolVIII(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the third  reactor line and reagents from bottle II.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(False)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()
            
def ReactorIX_SolIX(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the third  reactor line and reagents from bottle III.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(False)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()

def ReactorX_SolX(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the fourth  reactor line and reagents from bottle I.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()

def ReactorXI_SolXI(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
   """
   Executes the reaction using the fourth  reactor line and reagents from bottle II.

   Parameters:
       o (int): Experiment number.
       reac (int): Reactor identifier.
       reacTemp (float): Reactor temperature.
       wt1 (int): Waiting time.
       fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
       setpressure (float): Target pressure.

   Returns:
       None
   """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(True)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()
            
def ReactorXII_SolXII(o,reac,reacTemp,wt1,fr1,fr2,fr3,fr4,setpressure):
    """
    Executes the reaction using the fourth  reactor line and reagents from bottle III.

    Parameters:
        o (int): Experiment number.
        reac (int): Reactor identifier.
        reacTemp (float): Reactor temperature.
        wt1 (int): Waiting time.
        fr1, fr2, fr3, fr4 (float): Flow rates for pumps.
        setpressure (float): Target pressure.

    Returns:
        None
    """
    
    class Valve:
        def __init__(self, stateSetter) -> None:
            self.setter = stateSetter

        def setValveState(self, state):
            self.setter(state)
    client = rs.RSeriesClient('opc.tcp://localhost:43344')

    try:
        reac=str(reac)
        print("Reaction development, Experiment #: "+str(o) +"\n")
        
        isConnected = client.connect()
        client.getRSeries().getManualControl().setWCValveState(True)
        'Reactor Temperature'
        client.getRSeries().getManualControl().getR4I().getReactors()[reac].setTemperature(reacTemp)
        'Valves'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setSRValveState(False)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setSRValveState(True)    
        client.getRSeries().getManualControl().setWCValveState(True)
        'Flow rates'
        client.getRSeries().getManualControl().getModule(0).getPumpA().setFlowRate(fr1)
        client.getRSeries().getManualControl().getModule(0).getPumpB().setFlowRate(fr2)
        client.getRSeries().getManualControl().getModule(1).getPumpA().setFlowRate(fr3)
        client.getRSeries().getManualControl().getModule(1).getPumpB().setFlowRate(fr4)
        
        'Tiempo de Espera 3'   
        a="Reaction development"
        chronometer(wt1,a,o,setpressure)
    finally:
        if isConnected:
            client.disconnect()